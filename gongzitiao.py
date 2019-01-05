# -*- coding: UTF-8 -*-
import openpyxl
import smtplib
from email.mime.text import MIMEText


host = 'smtp.163.com'           # 设置发件服务器地址
port = 465                      # 设置发件服务器端口号
sender = 'xxxyyy@163.com'       # 设置发件邮箱
pwd = 'aaa20190105'             # 设置发件邮箱授权码
send_to_sub = 'x月工资条'


def get_html(person_data):
    html_content = """
<html>
    <head>
        <meta http-equiv="Content-Type" content="text/html; charset=utf-8" />
        <title>工资条</title>
    </head>
    <body>
        <div id="container">
            <p><strong>工资条：</strong></p>
            <div id="content">
                <table width="500" border="2">
                    <tr>
                        <td><strong>姓名</strong></td>
                        <td><strong>月份</strong></td>
                        <td><strong>部门</strong></td>
                        <td><strong>工号</strong></td>
                        <td><strong>基本工资</strong></td>
                        <td><strong>岗位工资</strong></td>
                        <td><strong>绩效工资</strong></td>
                        <td><strong>税前</strong></td>
                    </tr>
                    <tr>
                        <td>""" + person_data[0] + """</td>
                        <td>""" + person_data[1] + """</td>
                        <td>""" + person_data[2] + """</td>
                        <td>""" + person_data[3] + """</td>
                        <td>""" + person_data[4] + """</td>
                        <td>""" + person_data[5] + """</td>
                        <td>""" + person_data[6] + """</td>
                        <td>""" + person_data[7] + """</td>
                    </tr>
                </table>
            </div>
        </div>
    </body>
</html>
"""
    return html_content


def get_data():
    wb = openpyxl.load_workbook('gongzi.xlsx')
    ws = wb['Sheet1']
    all_data = []
    for row in range(2, ws.max_row + 1):
        per_list = []
        for cell in ws[row]:
            per_list.append(str(cell.value))
        all_data.append(per_list)
    return all_data


def send_mail(to_list, sub, content):
    me = "<"+sender+">"
    msg = MIMEText(content, _subtype='html', _charset='utf-8')
    msg['Subject'] = sub
    msg['From'] = me
    msg['To'] = ";".join(to_list)
    try:
        server = smtplib.SMTP()
        server.connect(host)
        server.login(sender, pwd)
        server.sendmail(me, to_list, msg.as_string())
        server.close()
        return True
    except smtplib.SMTPException as e:
        print(str(e))
        return False

if __name__ == '__main__':
    datas = get_data()          # 获取excel数据，转换为html
    for data in datas:
        send_to_list = [data[-1]]
        send_to_html = get_html(data)
        if send_mail(send_to_list, send_to_sub, send_to_html):
            print("发送成功")
        else:
            print("发送失败")
